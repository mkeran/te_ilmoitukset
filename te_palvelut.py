import csv
import logging
import xml.etree.ElementTree as ET
import urllib.request
import datetime as dt
import openpyxl.worksheet.worksheet as ws
import openpyxl
import xlsxwriter
from openpyxl.styles import Font
from os import path

###Versio, joka hakee tiedot ja lisää edelliseen tiedostoon uudet ilmoitukset

logger = logging.getLogger(__name__)


def check_item(item):
    del_titles = get_del_titles()
    last_load_time = get_last_time_on_file()

    pub_date_text = item.find("pubDate").text
    if before_time_cut(last_load_time, pub_date_text):
        return None

    title: str = item.find("title").text
    if title.startswith("Lisää ilmoituksia"):
        return None

    title_list = title.split(",")  # 0=Työn nimi, -1 = paikkakunta, kaikki välillä oleva on höttöä
    for x in del_titles:
        if len(x) < 3:
            continue
        if x in title_list[0].lower():
            return None

    data = {}
    data["Otsikko"] = title_list[0]
    data["Linkki"] = item.find("link").text
    data["Lisätietoja"] = ",".join(title_list[1:-1])
    data["Paikkakunta"] = title_list[-1]
    data["Julkaistu"] = dt.datetime.strptime(pub_date_text, '%a, %d %b %Y %H:%M:%S %z').strftime(
        '%Y %m %d %H:%M:%S')
    return data


def check_pubdates(channel):
    last_load_time = get_last_time_on_file()
    time_before_latest = False
    for item in channel.findall("item"):
        pub_date_text = item.find("pubDate").text
        if before_time_cut(last_load_time, pub_date_text):
            time_before_latest = True

    if time_before_latest == False:
        logger.info("Dataa haettu liian harvoin")

    newtime = channel.findall("item")[0].find("pubDate").text
    return newtime


def xml_file_to_list(xml: bin):
    tree = ET.parse(xml)
    root = tree.getroot()

    results = []
    count_skipped_items = 0

    for channel in root:
        newtime = check_pubdates(channel)
        for item in channel.findall("item"):
            data = check_item(item)
            if data == None:
                count_skipped_items += 1
                continue
            results.append(data)

    logger.info(f"Uusien lisättävien ilmoitusten määrä: {len(results)}")
    logger.info(f"Hylättyjen ilmoitusten määrä: {count_skipped_items}")

    return results, newtime


def get_last_time_on_file():
    with open("last_time_obj.txt") as f:
        time = f.read()
        # time_obj = dt.datetime.strptime(time, '%a, %d %b %Y %H:%M:%S %z').replace(tzinfo=None)
        return time


def set_last_time_on_file(time:str):
    with open("last_time_obj.txt", "w") as f:
        f.write(time)


def before_time_cut(time_cut:str, time:str):
    time_cut_obj = dt.datetime.strptime(time_cut, '%a, %d %b %Y %H:%M:%S %z').replace(tzinfo=None)
    time_obj = dt.datetime.strptime(time, '%a, %d %b %Y %H:%M:%S %z').replace(tzinfo=None)
    if time_obj < time_cut_obj:
        return True
    return False


def get_del_titles():
    with open("del_titles.txt", encoding="UTF-8") as f:
        list = f.read().split("\n")
    return list


def create_excel(data:list, file:str):
    logger.info("Luodaan tiedostoa: %s", file)

    workbook = xlsxwriter.Workbook(file)
    worksheet = workbook.add_worksheet(name = "Sheet1")
    cell_format = workbook.add_format()
    cell_format.set_bold()

    worksheet.write_row(0, 0, data[0].keys(), cell_format=cell_format)

    worksheet.set_column(0, 2, 50)
    worksheet.set_column(3, 4, 30)

    workbook.close()


def add_list_to_excel(data: list, file: str):
    if not path.isfile(file):
        create_excel(data, file)
    logger.debug("Lisätään tiedostoon: %s", file)

    workbook = openpyxl.load_workbook(file)
    worksheet: ws.Worksheet = workbook["Sheet1"]
    row = worksheet.max_row + 1

    logger.debug("Rivien määrä: %s", str(row + len(data) - 1))

    #luku loppuu, kirjoitus alkaa

    for dict in data[::-1]:
        worksheet.append(list(dict.values()))
        cell = worksheet.cell(row, 2)
        cell.hyperlink = cell.value
        cell.font = Font(underline='single', color='0563C1')
        row += 1
    workbook.save(file)
    workbook.close()


def clear_excel(file: str):
    workbook = openpyxl.load_workbook(file)
    worksheet: ws.Worksheet = workbook["Sheet1"]

    worksheet.delete_rows(2, worksheet.max_row)

    workbook.save(file)
    workbook.close()

def delete_nrows_from_excel(file:str, nrows: int):
    workbook = openpyxl.load_workbook(file)
    worksheet: ws.Worksheet = workbook["Sheet1"]

    worksheet.delete_rows(2, nrows)

    workbook.save(file)
    workbook.close()


def excel_too_full(file:str):
    workbook = openpyxl.load_workbook(file)
    worksheet: ws.Worksheet = workbook["Sheet1"]
    bool = False
    if worksheet.max_row > 2000:
        bool = True
    workbook.close()
    return bool


def main():
    # logger.info("ollaan mainissa")
    url = "https://paikat.te-palvelut.fi/tpt-api/tyopaikat.rss?alueet=Helsinki,Vantaa,Kerava&ilmoitettuPvm=3&vuokrapaikka=---"
    excel_file_name = "te_palvelut_excel.xlsx"
    excel_file = path.abspath(excel_file_name)
    if path.isfile(excel_file):
        if excel_too_full(excel_file):
            delete_nrows_from_excel(excel_file, 1000)
    logger.info(f"haetaan tiedot {get_last_time_on_file()} lähtien")
    xml = urllib.request.urlopen(url)

    data, newtime = xml_file_to_list(xml)
    try:
        add_list_to_excel(data, excel_file)
    except Exception:
        logger.info("Tietojen vienti exceliin epäonnistui")
    else:
        print("Uusi aika asetettu")
        set_last_time_on_file(newtime)


if __name__ == '__main__':
    # clear_excel("te_palvelut_excel.xlsx")
    main()